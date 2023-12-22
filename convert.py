from menu2_ui import Ui_Menu
import os
from pathlib import Path
import xlwings as xw
from PyPDF2 import PdfMerger, PdfReader, PdfWriter
from docx2pdf import convert
import base64
from PIL import Image
from reportlab.pdfgen import canvas
import fitz  # PyMuPDF
from pdf2docx import Converter
import pdfplumber
import openpyxl

def excel_to_pdf(self: Ui_Menu, file_path):
    with xw.App() as app:
        app.visible = False
        book = app.books.open(file_path)
        pdf_paths = []  # Lista para armazenar os caminhos dos arquivos PDF
        output_path = os.path.join(os.path.dirname(file_path), f"{os.path.splitext(os.path.basename(file_path))[0]}_sheet.pdf")
        for sheet_index, sheet in enumerate(book.sheets):
            # Construir o caminho para o arquivo PDF
            pdf_file = os.path.join(os.path.dirname(file_path), f"{os.path.splitext(os.path.basename(file_path))[0]}_sheet{sheet_index}.pdf")
            pdf_path = Path(pdf_file)

            # Salvar a planilha atual como um arquivo PDF
            sheet.to_pdf(path=pdf_path, show=False)

            # Adicionar o caminho do arquivo PDF à lista
            pdf_paths.append(pdf_path)

        merger = PdfMerger()

        for pdf in pdf_paths:
            merger.append(pdf)


        merger.write(output_path)
        merger.close()

        # code_base64 = pdf_to_base64(output_path)

        for pdf in pdf_paths:
            os.remove(pdf)
        # os.remove(output_path)

    return output_path


    
def word_to_pdf(self: Ui_Menu , file_path):
    pdf_file = os.path.join(os.path.dirname(file_path), f"{os.path.splitext(os.path.basename(file_path))[0]}.pdf")
    convert(file_path , pdf_file)
    # code_base64 = pdf_to_base64(pdf_file)
    # os.remove(pdf_file)
    return pdf_file


def pdf_to_base64(pdf_path, page_number):
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PdfReader(pdf_file)

        page = pdf_reader.pages[page_number]

        # Converte a página em um novo arquivo PDF temporário
        temp_pdf_path = os.path.join(os.path.dirname(pdf_path), f"{os.path.splitext(os.path.basename(pdf_path))[0]}temp_file_pdf_export.pdf")
        temp_pdf_writer = PdfWriter()
        temp_pdf_writer.add_page(page)

        with open(temp_pdf_path, 'wb') as temp_pdf_file:
            temp_pdf_writer.write(temp_pdf_file)

        # Lê os bytes do novo arquivo PDF temporário
        with open(temp_pdf_path, 'rb') as temp_pdf_file:
            pdf_bytes = temp_pdf_file.read()

        # Codifica os bytes para base64
        base64_page = base64.b64encode(pdf_bytes)
        os.remove(temp_pdf_path)
        return base64_page



def base64_to_pdf(codigo_path, file_path):
    # with open(codigo_path, 'rb') as codigo_file:
    #     # Lê o código base64 do arquivo
    codigo_base64 = codigo_path
    local_file = os.path.join(os.path.dirname(file_path), f"{os.path.splitext(os.path.basename(file_path))[0]}temp_new_file.pdf")
    
    # Decodifica o código base64 para bytes
    pdf_bytes = base64.b64decode(codigo_base64)
    
    # Cria um novo arquivo PDF
    with open(local_file, 'wb') as new_pdf_file:
        new_pdf_file.write(pdf_bytes)
    
    return local_file
        

def img_to_pdf(self: Ui_Menu, file_path):
    imagem = Image.open(file_path)
    pdf_file = os.path.join(os.path.dirname(file_path), f"{os.path.splitext(os.path.basename(file_path))[0]}.pdf")
    largura, altura = imagem.size

    # Criação de um objeto PDF
    pdf = canvas.Canvas(pdf_file, pagesize=(largura, altura))


    # Adiciona a imagem ao PDF
    pdf.drawInlineImage(file_path, 0, 0, largura, altura)

    # Salva o PDF
    pdf.save()
    
    # code_base64 = pdf_to_base64(pdf_file)
    # os.remove(pdf_file)
    return pdf_file



def scale_image(self: Ui_Menu, file_path, max_width, max_height):
    try:
        # Defina o tamanho máximo em pixels para uma página A4 (210mm x 297mm a 300 DPI)
        img = Image.open(file_path)
        width, height = img.size

        # Verifique se a imagem precisa ser redimensionada
        if width > max_width or height > max_height:
            # Calcule a nova escala mantendo a proporção
            scale = min(max_width / width, max_height / height)
            new_width = int(width * scale)
            new_height = int(height * scale)

            # Redimensione a imagem
            # img = img.resize((new_width, new_height), Image.ANTIALIAS)
            img = img.resize((new_width, new_height), Image.ANTIALIAS if hasattr(Image, 'ANTIALIAS') else 3)
            name_img = os.path.join(os.path.dirname(file_path), f"{os.path.splitext(os.path.basename(file_path))[0]}_SCALE_.png")
            # Salve a nova imagem redimensionada
            # scaled_path = f"{file_path}_scaled_"
            scaled_path = name_img
            img.save(scaled_path)

            return scaled_path
        else:
            return False
    except Exception as e:
        print(f"Error scaling image: {e}")
        return None
    


def pdf_to_word(self: Ui_Menu, pdf_file , local_save_pg):
    cv = Converter(pdf_file)
    docx_file = os.path.join(os.path.dirname(local_save_pg), f"{os.path.splitext(os.path.basename(local_save_pg))[0]}.docx")
    print(docx_file)
    cv.convert(docx_file, start=0, end=None)
    cv.close()



def pdf_to_excel(self: Ui_Menu, pdf_file, local_save_pg):

    # Inicialize uma lista para armazenar as coordenadas de cada palavra
    coordenadas_palavras = []

    # Inicialize um conjunto para armazenar as coordenadas y0
    coordenadas_y0 = set()

    # Inicialize um dicionário para armazenar as palavras por linha
    palavras_por_linha = {}

    # Inicialize uma variável para rastrear a planilha atual
    sheet_index = 1

    # Abra o arquivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Sheet {sheet_index}'

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            lista = []
            colunas = 0
            valor = 0
            lista_texto = []
            lista_valores = []
            dict_distancia = {}
            lista_novos_valores = []
            contador = 0
            coll = 1
            coordenadas_palavras = []
            for word in page.extract_words():
                texto = word['text']  # Obtém o texto da palavra
                bbox = word['x0'], word['top'], word['x1'], word['bottom']  # Obtém as coordenadas da caixa delimitadora
                coordenadas_palavras.append((texto, bbox))
                y0 = word['top']  # Obtém a coordenada y0
                coordenadas_y0.add(y0)  # Adiciona a coordenada y0 ao conjunto

            # O número total de linhas é igual ao tamanho do conjunto de coordenadas y0
            numero_de_linhas = len(coordenadas_y0)

            # Agora, você pode transformar as coordenadas y0 em números de linha
            coordenadas_y0_ordenadas = sorted(coordenadas_y0)  # Ordene as coordenadas
            coordenadas_y0_para_linhas = {coord: linha for linha, coord in enumerate(coordenadas_y0_ordenadas, start=1)}

            # Inicialize a lista de palavras por linha
            for linha in range(1, numero_de_linhas + 1):
                palavras_por_linha[linha] = []

            # Exibe as coordenadas de cada palavra, seu número de linha e distância para a próxima palavra
            for texto, bbox in coordenadas_palavras:
                y0 = bbox[1]  # Obtém a coordenada y0 da caixa delimitadora
                linha = coordenadas_y0_para_linhas[y0]  # Obtém o número da linha

                # Calcula a distância para a próxima palavra na mesma linha
                distancias = []
                for proxima_palavra, proximo_bbox in coordenadas_palavras:
                    proxima_y0 = proximo_bbox[1]
                    if coordenadas_y0_para_linhas[proxima_y0] == linha:
                        distancia = proximo_bbox[0] - bbox[2]
                        distancias.append(distancia)

                # Adiciona a palavra à lista de palavras por linha
                palavras_por_linha[linha].append((texto, bbox, distancias))

            # Verifica se a página mudou
            if page.page_number != sheet_index:
                sheet_index = page.page_number
                sheet = workbook.create_sheet(title=f'Sheet {sheet_index}')

            # Exibe as palavras agrupadas por linha com as posições atuais
            for linha in range(1, numero_de_linhas + 1):
                lista = []
                colunas = 0
                valor = 0
                lista_texto = []
                lista_valores = []
                dict_distancia = {}
                lista_novos_valores = []
                contador = 0
                coll = 1
        
                for texto, bbox, _ in palavras_por_linha[linha]:
                    x0, y0, x1, y1 = bbox
                    lista_texto.append(texto)
                    if len(lista) == 0:
                        lista.append(x1)
                        chave = f"chave_{contador}"
                        dict_distancia[chave] = {
                            'texto': texto,
                            'distancia_palavra_anterior': None
                        }
                        contador += 1
                    else:
                        chave = f"chave_{contador}"
                        valor = int(x0 - lista[-1])
                        dict_distancia[chave] = {
                            'texto': texto,
                            'distancia_palavra_anterior': valor
                        }
                        lista_valores.append(valor)
                        lista.append(x1)
                        contador += 1
                if len(lista_valores) == 0:
                    colunas = 1
                else:
                    menor_numero = min(lista_valores)
                    vezes = 0
                    for n_vezes in lista_valores:
                        if n_vezes == menor_numero:
                            vezes += 1
                    if vezes == 1:
                        vezes = 0
                        colunas = len(lista_texto) - vezes
                        for item in lista_texto:
                            lista_novos_valores.append(item)
                    else:
                        colunas = len(lista_texto) - vezes
                        for chave, palavra in dict_distancia.items():
                            dist_dicionario = palavra.get('distancia_palavra_anterior')
                            text_dict = palavra.get('texto')
                            if dist_dicionario == menor_numero:
                                lista_novos_valores[-1] = (f'{lista_novos_valores[-1]} {text_dict}')
                            else:
                                lista_novos_valores.append(text_dict)

                # Preencha o arquivo Excel com os dados
                for palavra in lista_novos_valores:
                    sheet.cell(row=linha, column=coll, value=palavra)
                    coll += 1

    # Salvar a planilha Excel
    xlsx_file =os.path.join(os.path.dirname(local_save_pg), f"{os.path.splitext(os.path.basename(local_save_pg))[0]}.xlsx")
    
    # Salvar a planilha Excel
    workbook.save(xlsx_file)
    workbook.close()


def pdf_to_png(self:Ui_Menu, pdf_file, local_save_img):
    with open(pdf_file, 'rb') as pdf_path:
        pdf_reader = fitz.open(pdf_path)
        
        for n_page in range(pdf_reader.page_count):
            page = pdf_reader.load_page(n_page)

            # Ajuste a resolução (dpi) da imagem exportada para 600 DPI
            image = page.get_pixmap(matrix=fitz.Matrix(600 / 72, 600 / 72))

            # Nome do arquivo de imagem de saída
            output_image = os.path.join(
                os.path.dirname(local_save_img),
                f"{os.path.splitext(os.path.basename(local_save_img))[0]}_page_{n_page}.png"
            )

            # Salve a imagem no mesmo local selecionado
            image.save(output_image, "png")


def pdf_to_jpg(self:Ui_Menu, pdf_file, local_save_img):
    doc = fitz.open(pdf_file)

    for pagina_numero in range(doc.page_count):
        pagina = doc[pagina_numero]
        magem = pagina.get_pixmap()
        pixel = magem.irect
        imagem_pillow = Image.frombytes("RGB", (pixel.x1, pixel.y1), magem.samples)

        output_image = os.path.join(
                os.path.dirname(local_save_img),
                f"{os.path.splitext(os.path.basename(local_save_img))[0]}_page_{pagina_numero}.jpeg"
            )
        imagem_pillow.save(output_image, "JPEG")
        # imagem_pillow.save(local_save_img, "JPEG")

    doc.close()