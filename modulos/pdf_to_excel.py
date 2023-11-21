# from menu2_ui import Ui_Menu
# from message import show_success_message
# import os
# from io import StringIO
# from pdfminer.high_level import extract_text_to_fp
# from pdfminer.layout import LAParams
# from bs4 import BeautifulSoup
# import openpyxl
# from openpyxl.styles import Font, Alignment, NamedStyle ,Border, Side
# from openpyxl.utils import get_column_letter
# output_string = StringIO()
# from openpyxl import Workbook
# from openpyxl.styles import NamedStyle
# from openpyxl.styles.borders import Border
# from openpyxl.styles import Side

# Crie o estilo personalizado uma vez
# workbook = openpyxl.Workbook()
# custom_style = NamedStyle(name='custom_style')

# def apply_custom_style(cell, style_dict):
#     # Definir o estilo de borda da célula
#     border = Border(
#         left=Side(border_style=style_dict.get('border-left-style', 'none'),
#                   color=style_dict.get('border-left-color', '000000')),
#         right=Side(border_style=style_dict.get('border-right-style', 'none'),
#                    color=style_dict.get('border-right-color', '000000')),
#         top=Side(border_style=style_dict.get('border-top-style', 'none'),
#                  color=style_dict.get('border-top-color', '000000')),
#         bottom=Side(border_style=style_dict.get('border-bottom-style', 'none'),
#                     color=style_dict.get('border-bottom-color', '000000'))
#     )

#     cell.border = border

# def convert_pixel_to_characters(pixel_value):
#     return max(1, int(pixel_value) // 10)

# def pdf_to_excel(main_widget: Ui_Menu, pdf_file, local_save_pg):
#      # Especificar o caminho para o arquivo HTML
#     html_file = os.path.join(os.path.dirname(local_save_pg), f"{os.path.splitext(os.path.basename(local_save_pg))[0]}.html")
    
#     # Abrir o arquivo PDF e extrair o texto para o arquivo HTML
#     with open(pdf_file, 'rb') as fin:
#         extract_text_to_fp(fin, output_string, laparams=LAParams(), output_type='html', codec=None)
    
#     with open(html_file, 'w', encoding='utf-8') as html_out:
#         html_out.write(output_string.getvalue())

#     # Ler o arquivo HTML
#     with open(html_file, "r", encoding="utf-8") as file:
#         html_content = file.read()

#     # Crie um novo arquivo Excel
#     docx_file = os.path.join(os.path.dirname(local_save_pg), f"{os.path.splitext(os.path.basename(local_save_pg))[0]}.xlsx")
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     current_row = 1
#     cell_height = 15  # Altura de cada célula

#     # Analisar o HTML com BeautifulSoup
#     soup = BeautifulSoup(html_content, 'html.parser')

#     # Encontrar todas as tags <div> com informações de células
#     divs = soup.find_all('div')

#     for div in divs:
#         # Extrair o texto
#         cell_text = div.get_text()

#         # Aplicar o estilo à célula
#         style = div.get('style', '')
#         styles = style.split(';')
#         style_dict = {}
#         for s in styles:
#             if ':' in s:
#                 key, value = s.split(':')
#                 style_dict[key.strip()] = value.strip()

#         # Calcular as coordenadas da célula com base no estilo
#         left = float(style_dict.get('left', '0px').replace('px', '').strip() or '0')
#         top = float(style_dict.get('top', '0px').replace('px', '').strip() or '0')

#         # Calcular a largura da coluna com base no valor width e definir a largura da coluna
#         cell_width = float(style_dict.get('width', '75px').replace('px', '').strip() or '75')
#         cell_width_characters = convert_pixel_to_characters(cell_width)
#         column = int(left // cell_width) + 1
#         sheet.column_dimensions[get_column_letter(column)].width = cell_width_characters

#         # Verificar se a linha atual já foi preenchida
#         while current_row < top // cell_height:
#             current_row += 1

#         # Aplicar o texto na célula apropriada com estilo personalizado
#         cell = sheet.cell(row=current_row, column=column)
#         apply_custom_style(cell, style_dict)
#         cell.value = cell_text

    # # Salvar a planilha Excel
    #         workbook.save(xlsx_file)
    #         os.remove(html_file)
    #         print('Planilha criada com sucesso.')
    #         output_string.truncate(0)
    #         show_success_message(main_widget)
###########################################################################################################3

# import pdfplumber
# import openpyxl
# import os
# from menu2_ui import Ui_Menu
# from message import show_success_message

# def pdf_to_excel(main_widget: Ui_Menu, pdf_file , local_save_pg):

#     # Inicialize uma lista para armazenar as coordenadas de cada palavra
#     coordenadas_palavras = []

#     # Inicialize um conjunto para armazenar as coordenadas y0
#     coordenadas_y0 = set()

#     # Inicialize um dicionário para armazenar as palavras por linha
#     palavras_por_linha = {}

#     with pdfplumber.open(pdf_file) as pdf:
#         for page in pdf.pages:
#             for word in page.extract_words():
#                 texto = word['text']  # Obtém o texto da palavra
#                 bbox = word['x0'], word['top'], word['x1'], word['bottom']  # Obtém as coordenadas da caixa delimitadora
#                 coordenadas_palavras.append((texto, bbox))
#                 y0 = word['top']  # Obtém a coordenada y0
#                 coordenadas_y0.add(y0)  # Adiciona a coordenada y0 ao conjunto

#         # O número total de linhas é igual ao tamanho do conjunto de coordenadas y0
#         numero_de_linhas = len(coordenadas_y0)

#         # Agora, você pode transformar as coordenadas y0 em números de linha
#         coordenadas_y0_ordenadas = sorted(coordenadas_y0)  # Ordene as coordenadas
#         coordenadas_y0_para_linhas = {coord: linha for linha, coord in enumerate(coordenadas_y0_ordenadas, start=1)}

#         # Inicialize a lista de palavras por linha
#         for linha in range(1, numero_de_linhas + 1):
#             palavras_por_linha[linha] = []

#         # Exibe as coordenadas de cada palavra, seu número de linha e distância para a próxima palavra
#         for texto, bbox in coordenadas_palavras:
#             y0 = bbox[1]  # Obtém a coordenada y0 da caixa delimitadora
#             linha = coordenadas_y0_para_linhas[y0]  # Obtém o número da linha

#             # Calcula a distância para a próxima palavra na mesma linha
#             distancias = []
#             for proxima_palavra, proximo_bbox in coordenadas_palavras:
#                 proxima_y0 = proximo_bbox[1]
#                 if coordenadas_y0_para_linhas[proxima_y0] == linha:
#                     distancia = proximo_bbox[0] - bbox[2]
#                     distancias.append(distancia)

#             # Adiciona a palavra à lista de palavras por linha
#             palavras_por_linha[linha].append((texto, bbox, distancias))

#         # Exibe as palavras agrupadas por linha
#         for linha in range(1, numero_de_linhas + 1):
#             # print(f"Linha {linha}:")
#             for texto, _, distancias in palavras_por_linha[linha]:
#                 # print(f"Texto: {texto} (Distâncias para a próxima palavra: {distancias})")
#                 pass


#         # Crie um novo arquivo Excel
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active



#             # Exibe as palavras agrupadas por linha com as posições atuais
#         for linha in range(1, numero_de_linhas + 1):
#             lista = []
#             colunas = 0
#             valor = 0
#             lista_texto = []
#             lista_valores = []
#             dict_distancia = {}
#             lista_novos_valores = []
#             contador = 0
#             coll = 1
        
#             # print(f"Linha {linha}:")
#             for texto, bbox, _ in palavras_por_linha[linha]:
#                 x0, y0, x1, y1 = bbox
#                 # print(f"Texto: {texto}, x1={x1})")
#                 lista_texto.append(texto)
#                 if len(lista) == 0:
#                     lista.append(x1)
#                     chave = f"chave_{contador}"
#                     dict_distancia[chave] = {
#                         'texto':texto,
#                         'distancia_palavra_anterior': None
#                     }
#                     contador += 1
#                 else:
#                     chave = f"chave_{contador}"
#                     valor = int(x0 - lista[-1])
#                     dict_distancia[chave] = {
#                         'texto':texto,
#                         'distancia_palavra_anterior':valor
#                     }
#                     lista_valores.append(valor)
#                     lista.append(x1)
#                     contador += 1
#             if len(lista_valores) == 0:
#                 colunas = 1

#             else:
#                 menor_numero = min(lista_valores)
#                 vezes = 0
#                 for n_vezes in lista_valores:
#                     if n_vezes == menor_numero:
#                         vezes += 1
#                 if vezes == 1:
#                     vezes = 0
#                     colunas = len(lista_texto) - vezes
#                     for item in lista_texto:
#                         lista_novos_valores.append(item)
#                 else:
#                     colunas = len(lista_texto) - vezes
#                     for chave, palavra in dict_distancia.items():
#                         dist_dicionario = palavra.get('distancia_palavra_anterior')
#                         text_dict = palavra.get('texto')
#                         if dist_dicionario == menor_numero:
#                             lista_novos_valores[-1] = (f'{lista_novos_valores[-1]} {text_dict}')
#                         else:
#                             lista_novos_valores.append(text_dict)



                        
#             # print(f'Linha {linha} {lista_novos_valores} - Número Estimado de Colunas: {colunas}')


#             # Preencha o arquivo Excel com os dados
#             for palavra in lista_novos_valores:
#                 sheet.cell(row=linha, column=coll, value=palavra)
#                 coll += 1

#         xlsx_file =os.path.join(os.path.dirname(local_save_pg), f"{os.path.splitext(os.path.basename(local_save_pg))[0]}.xlsx")

#     # Salvar a planilha Excel
#     workbook.save(xlsx_file)
#     print('Planilha criada com sucesso.')
#     show_success_message(main_widget)


##############################################################################################################################

import pdfplumber
import openpyxl
import os
from menu2_ui import Ui_Menu
from message import show_success_message

def pdf_to_excel(main_widget: Ui_Menu, pdf_file, local_save_pg):

    # Inicialize uma lista para armazenar as coordenadas de cada palavra
    coordenadas_palavras = []

    # Inicialize um conjunto para armazenar as coordenadas y0
    coordenadas_y0 = set()

    # Inicialize um dicionário para armazenar as palavras por linha
    palavras_por_linha = {}

    # Inicialize uma variável para rastrear a planilha atual
    sheet_index = 1

    # Abra o arquivo Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f'Sheet {sheet_index}'

    with pdfplumber.open(pdf_file) as pdf:
        for page in pdf.pages:
            lista = []
            colunas = 0
            valor = 0
            lista_texto = []
            lista_valores = []
            dict_distancia = {}
            lista_novos_valores = []
            contador = 0
            coll = 1
            coordenadas_palavras = []
            for word in page.extract_words():
                texto = word['text']  # Obtém o texto da palavra
                bbox = word['x0'], word['top'], word['x1'], word['bottom']  # Obtém as coordenadas da caixa delimitadora
                coordenadas_palavras.append((texto, bbox))
                y0 = word['top']  # Obtém a coordenada y0
                coordenadas_y0.add(y0)  # Adiciona a coordenada y0 ao conjunto

            # O número total de linhas é igual ao tamanho do conjunto de coordenadas y0
            numero_de_linhas = len(coordenadas_y0)

            # Agora, você pode transformar as coordenadas y0 em números de linha
            coordenadas_y0_ordenadas = sorted(coordenadas_y0)  # Ordene as coordenadas
            coordenadas_y0_para_linhas = {coord: linha for linha, coord in enumerate(coordenadas_y0_ordenadas, start=1)}

            # Inicialize a lista de palavras por linha
            for linha in range(1, numero_de_linhas + 1):
                palavras_por_linha[linha] = []

            # Exibe as coordenadas de cada palavra, seu número de linha e distância para a próxima palavra
            for texto, bbox in coordenadas_palavras:
                y0 = bbox[1]  # Obtém a coordenada y0 da caixa delimitadora
                linha = coordenadas_y0_para_linhas[y0]  # Obtém o número da linha

                # Calcula a distância para a próxima palavra na mesma linha
                distancias = []
                for proxima_palavra, proximo_bbox in coordenadas_palavras:
                    proxima_y0 = proximo_bbox[1]
                    if coordenadas_y0_para_linhas[proxima_y0] == linha:
                        distancia = proximo_bbox[0] - bbox[2]
                        distancias.append(distancia)

                # Adiciona a palavra à lista de palavras por linha
                palavras_por_linha[linha].append((texto, bbox, distancias))

            # Verifica se a página mudou
            if page.page_number != sheet_index:
                sheet_index = page.page_number
                sheet = workbook.create_sheet(title=f'Sheet {sheet_index}')

            # Exibe as palavras agrupadas por linha com as posições atuais
            for linha in range(1, numero_de_linhas + 1):
                lista = []
                colunas = 0
                valor = 0
                lista_texto = []
                lista_valores = []
                dict_distancia = {}
                lista_novos_valores = []
                contador = 0
                coll = 1
        
                for texto, bbox, _ in palavras_por_linha[linha]:
                    x0, y0, x1, y1 = bbox
                    lista_texto.append(texto)
                    if len(lista) == 0:
                        lista.append(x1)
                        chave = f"chave_{contador}"
                        dict_distancia[chave] = {
                            'texto': texto,
                            'distancia_palavra_anterior': None
                        }
                        contador += 1
                    else:
                        chave = f"chave_{contador}"
                        valor = int(x0 - lista[-1])
                        dict_distancia[chave] = {
                            'texto': texto,
                            'distancia_palavra_anterior': valor
                        }
                        lista_valores.append(valor)
                        lista.append(x1)
                        contador += 1
                if len(lista_valores) == 0:
                    colunas = 1
                else:
                    menor_numero = min(lista_valores)
                    vezes = 0
                    for n_vezes in lista_valores:
                        if n_vezes == menor_numero:
                            vezes += 1
                    if vezes == 1:
                        vezes = 0
                        colunas = len(lista_texto) - vezes
                        for item in lista_texto:
                            lista_novos_valores.append(item)
                    else:
                        colunas = len(lista_texto) - vezes
                        for chave, palavra in dict_distancia.items():
                            dist_dicionario = palavra.get('distancia_palavra_anterior')
                            text_dict = palavra.get('texto')
                            if dist_dicionario == menor_numero:
                                lista_novos_valores[-1] = (f'{lista_novos_valores[-1]} {text_dict}')
                            else:
                                lista_novos_valores.append(text_dict)

                # Preencha o arquivo Excel com os dados
                for palavra in lista_novos_valores:
                    sheet.cell(row=linha, column=coll, value=palavra)
                    coll += 1

    # Salvar a planilha Excel
    xlsx_file =os.path.join(os.path.dirname(local_save_pg), f"{os.path.splitext(os.path.basename(local_save_pg))[0]}.xlsx")
    
    # Salvar a planilha Excel
    workbook.save(xlsx_file)
    print('Planilha criada com sucesso.')
    show_success_message(main_widget)